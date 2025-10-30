"""
Генератор КУДиР (Книга учета доходов и расходов) для УСН "доходы минус расходы"
"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from app.database import crud
from app.config import settings
from datetime import date
from sqlalchemy.ext.asyncio import AsyncSession
import os
import logging

logger = logging.getLogger(__name__)


async def generate_kudir(
    session: AsyncSession,
    year: int,
    quarter: int = None
) -> openpyxl.Workbook:
    """
    Генерация КУДиР (Книга учета доходов и расходов)
    для УСН "доходы минус расходы"

    Args:
        session: Сессия БД
        year: Год
        quarter: Квартал (1-4), если None - весь год

    Returns:
        openpyxl.Workbook объект
    """
    # Проверяем наличие шаблона
    template_path = os.path.join(settings.TEMPLATES_PATH, 'kudir_usn_income_expense.xlsx')

    if os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
    else:
        # Создаем новый файл
        wb = openpyxl.Workbook()

    # Раздел I - Доходы и расходы
    ws = wb.active
    ws.title = "Раздел I"

    # Определяем период
    if quarter:
        start_date = date(year, (quarter - 1) * 3 + 1, 1)
        import calendar
        last_month = quarter * 3
        last_day = calendar.monthrange(year, last_month)[1]
        end_date = date(year, last_month, last_day)
        period_text = f"{year} год, {quarter} квартал"
    else:
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        period_text = f"{year} год"

    # Стили
    header_font = Font(bold=True, size=14)
    table_header_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Титульная информация
    ws['A1'] = f"Раздел I. Доходы и расходы"
    ws['A1'].font = header_font
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = center_align

    ws['A2'] = f"ООО \"Лепта\" (ИНН {settings.COMPANY_INN})"
    ws.merge_cells('A2:E2')
    ws['A2'].alignment = center_align

    ws['A3'] = period_text
    ws['A3'].font = Font(bold=True, size=12)
    ws.merge_cells('A3:E3')
    ws['A3'].alignment = center_align

    # Шапка таблицы
    headers = [
        ("№", 5),
        ("Дата и номер\nпервичного\nдокумента", 20),
        ("Содержание операции", 40),
        ("Доходы,\nучитываемые\nпри исчислении\nналоговой базы", 15),
        ("Расходы,\nучитываемые\nпри исчислении\nналоговой базы", 15)
    ]

    row = 5
    for col, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = table_header_font
        cell.alignment = center_align
        cell.border = border
        # Устанавливаем ширину колонки
        ws.column_dimensions[get_column_letter(col)].width = width

    # Получаем транзакции
    transactions = await crud.get_transactions_by_period(
        session,
        start_date,
        end_date,
        confirmed_only=True
    )

    # Сортируем по дате
    transactions.sort(key=lambda t: (t.date, t.created_at))

    # Заполняем данные
    row = 6
    total_income = 0
    total_expense = 0

    for idx, t in enumerate(transactions, start=1):
        # Номер
        cell = ws.cell(row=row, column=1, value=idx)
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

        # Дата и номер документа
        doc_text = t.date.strftime('%d.%m.%Y')
        if t.document_number:
            doc_text += f"\n№ {t.document_number}"
        cell = ws.cell(row=row, column=2, value=doc_text)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True)

        # Содержание операции
        content = t.description or ""
        if t.counterparty:
            content = f"{t.counterparty}\n{content}" if content else t.counterparty
        if t.category:
            content += f"\n({t.category.name})"

        cell = ws.cell(row=row, column=3, value=content)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Доходы / Расходы
        if t.type == 'income':
            cell = ws.cell(row=row, column=4, value=float(t.amount))
            cell.number_format = '#,##0.00'
            total_income += t.amount
            ws.cell(row=row, column=5).value = ""
        else:
            ws.cell(row=row, column=4).value = ""
            # Расходы учитываем только если категория tax_deductible
            if t.category and t.category.tax_deductible:
                cell = ws.cell(row=row, column=5, value=float(t.amount))
                cell.number_format = '#,##0.00'
                total_expense += t.amount
            else:
                ws.cell(row=row, column=5).value = ""

        # Применяем border ко всем ячейкам
        for col in range(4, 6):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')

        row += 1

    # Итого
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws.cell(row=row, column=1, value="ИТОГО:")
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='right')
    cell.border = border

    cell = ws.cell(row=row, column=4, value=float(total_income))
    cell.font = Font(bold=True)
    cell.number_format = '#,##0.00'
    cell.border = border
    cell.alignment = Alignment(horizontal='right')

    cell = ws.cell(row=row, column=5, value=float(total_expense))
    cell.font = Font(bold=True)
    cell.number_format = '#,##0.00'
    cell.border = border
    cell.alignment = Alignment(horizontal='right')

    # База налогообложения
    row += 2
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws.cell(row=row, column=1, value="База налогообложения (доходы - расходы):")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='right')

    tax_base = max(total_income - total_expense, 0)
    cell = ws.cell(row=row, column=4, value=float(tax_base))
    cell.font = Font(bold=True, size=12)
    cell.number_format = '#,##0.00'
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cell.border = border

    # Подпись
    row += 4
    ws.cell(row=row, column=1, value=f"Директор ООО \"Лепта\"")
    ws.cell(row=row, column=3, value="__________________")
    row += 2
    ws.cell(row=row, column=1, value=f"Дата: {date.today().strftime('%d.%m.%Y')}")

    logger.info(f"KUDiR generated for {period_text}: income={total_income}, expense={total_expense}")

    return wb


async def generate_kudir_file(
    session: AsyncSession,
    year: int,
    quarter: int = None,
    output_path: str = None
) -> str:
    """
    Сгенерировать КУДиР и сохранить в файл

    Args:
        session: Сессия БД
        year: Год
        quarter: Квартал
        output_path: Путь для сохранения (если None - в /tmp)

    Returns:
        Путь к созданному файлу
    """
    wb = await generate_kudir(session, year, quarter)

    if not output_path:
        quarter_suffix = f"_q{quarter}" if quarter else ""
        filename = f"kudir_{year}{quarter_suffix}.xlsx"
        output_path = f"/tmp/{filename}"

    wb.save(output_path)
    logger.info(f"KUDiR file saved: {output_path}")

    return output_path
