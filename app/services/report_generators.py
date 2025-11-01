"""
Генераторы отчетов для ФНС и фондов
"""
import logging
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Dict, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

from ..config import settings

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Базовый класс для генераторов отчетов"""

    def __init__(self, output_dir: str = "/opt/accounting-bot/reports"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)


class RSVGenerator(ReportGenerator):
    """
    Генератор РСВ (Расчет страховых взносов)
    Упрощенная версия - в реальности нужен XML для ФНС
    """

    def generate(
        self,
        year: int,
        quarter: int,
        payrolls: List[Dict],
        tax_data: Dict
    ) -> str:
        """
        Генерация РСВ отчета

        Returns:
            Путь к сохраненному файлу
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "РСВ"

        # Заголовок
        ws['A1'] = f"РАСЧЕТ СТРАХОВЫХ ВЗНОСОВ"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"за {quarter} квартал {year} года"
        ws['A2'].font = Font(bold=True, size=12)

        # Данные организации
        row = 4
        ws[f'A{row}'] = f"Организация: {settings.COMPANY_NAME}"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = f"ИНН: {settings.COMPANY_INN}"
        row += 2

        # Таблица расчетов
        headers = ['ФИО', 'Начислено', 'ПФР (22%)', 'ОМС (5.1%)', 'ФСС (2.9%)', 'ВНиМ (0.2%)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = Border(
                bottom=Side(style='thin'),
                top=Side(style='thin'),
                left=Side(style='thin'),
                right=Side(style='thin')
            )

        row += 1

        # Данные по сотрудникам
        totals = {
            'gross': Decimal('0'),
            'pension': Decimal('0'),
            'medical': Decimal('0'),
            'social': Decimal('0'),
            'injury': Decimal('0')
        }

        for payroll in payrolls:
            gross = payroll.get('gross_salary', 0)
            contributions = payroll.get('contributions', {})

            ws.cell(row=row, column=1, value=payroll.get('employee_name', ''))
            ws.cell(row=row, column=2, value=float(gross))
            ws.cell(row=row, column=3, value=float(contributions.get('pension', 0)))
            ws.cell(row=row, column=4, value=float(contributions.get('medical', 0)))
            ws.cell(row=row, column=5, value=float(contributions.get('social', 0)))
            ws.cell(row=row, column=6, value=float(contributions.get('injury', 0)))

            totals['gross'] += Decimal(str(gross))
            totals['pension'] += Decimal(str(contributions.get('pension', 0)))
            totals['medical'] += Decimal(str(contributions.get('medical', 0)))
            totals['social'] += Decimal(str(contributions.get('social', 0)))
            totals['injury'] += Decimal(str(contributions.get('injury', 0)))

            row += 1

        # Итого
        row += 1
        ws.cell(row=row, column=1, value='ИТОГО:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=float(totals['gross'])).font = Font(bold=True)
        ws.cell(row=row, column=3, value=float(totals['pension'])).font = Font(bold=True)
        ws.cell(row=row, column=4, value=float(totals['medical'])).font = Font(bold=True)
        ws.cell(row=row, column=5, value=float(totals['social'])).font = Font(bold=True)
        ws.cell(row=row, column=6, value=float(totals['injury'])).font = Font(bold=True)

        # Сумма к уплате
        total_contributions = totals['pension'] + totals['medical'] + totals['social'] + totals['injury']
        row += 2
        ws.cell(row=row, column=1, value='ВСЕГО К УПЛАТЕ:')
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2, value=float(total_contributions))
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)

        # Сохранение
        filename = f"RSV_{year}_Q{quarter}_{date.today().isoformat()}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)

        logger.info(f"RSV report generated: {filepath}")
        return str(filepath)


class SZVMGenerator(ReportGenerator):
    """
    Генератор СЗВ-М (Сведения о застрахованных лицах)
    """

    def generate(
        self,
        year: int,
        month: int,
        employees: List[Dict]
    ) -> str:
        """
        Генерация СЗВ-М

        Returns:
            Путь к сохраненному файлу
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "СЗВ-М"

        # Заголовок
        ws['A1'] = "СВЕДЕНИЯ О ЗАСТРАХОВАННЫХ ЛИЦАХ"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"СЗВ-М за {month:02d}.{year}"
        ws['A2'].font = Font(bold=True, size=12)

        row = 4
        ws[f'A{row}'] = f"Страхователь: {settings.COMPANY_NAME}"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        ws[f'A{row}'] = f"ИНН: {settings.COMPANY_INN}"
        row += 2

        # Заголовки таблицы
        headers = ['№', 'ФИО', 'СНИЛС', 'ИНН']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)

        row += 1

        # Данные
        for idx, employee in enumerate(employees, start=1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=employee.get('full_name', ''))
            ws.cell(row=row, column=3, value=employee.get('snils', ''))
            ws.cell(row=row, column=4, value=employee.get('inn', ''))
            row += 1

        # Подпись
        row += 2
        ws.cell(row=row, column=1, value='Руководитель: _____________')

        # Сохранение
        filename = f"SZV-M_{year}_{month:02d}_{date.today().isoformat()}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)

        logger.info(f"SZV-M report generated: {filepath}")
        return str(filepath)


class EFS1Generator(ReportGenerator):
    """
    Генератор ЕФС-1 (Единая форма сведений)
    """

    def generate(
        self,
        year: int,
        quarter: int,
        employees: List[Dict],
        contracts: List[Dict]
    ) -> str:
        """
        Генерация ЕФС-1

        Returns:
            Путь к сохраненному файлу
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "ЕФС-1"

        # Заголовок
        ws['A1'] = "ЕДИНАЯ ФОРМА СВЕДЕНИЙ (ЕФС-1)"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"за {quarter} квартал {year} года"
        ws['A2'].font = Font(bold=True, size=12)

        row = 4
        ws[f'A{row}'] = f"Организация: {settings.COMPANY_NAME}"
        row += 1
        ws[f'A{row}'] = f"ИНН: {settings.COMPANY_INN}"
        row += 2

        # Раздел 1: Сведения о трудовой деятельности
        ws.cell(row=row, column=1, value="Раздел 1. Сведения о трудовой деятельности")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 2

        headers = ['ФИО', 'Должность', 'Дата приема', 'Тип договора']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=header).font = Font(bold=True)

        row += 1

        for employee in employees:
            ws.cell(row=row, column=1, value=employee.get('full_name', ''))
            ws.cell(row=row, column=2, value=employee.get('position', 'Администратор'))
            if employee.get('hire_date'):
                ws.cell(row=row, column=3, value=employee['hire_date'].strftime('%d.%m.%Y'))
            ws.cell(row=row, column=4, value=employee.get('employment_type', ''))
            row += 1

        # Сохранение
        filename = f"EFS-1_{year}_Q{quarter}_{date.today().isoformat()}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)

        logger.info(f"EFS-1 report generated: {filepath}")
        return str(filepath)


class USNDeclarationGenerator(ReportGenerator):
    """
    Генератор декларации УСН
    """

    def generate(
        self,
        year: int,
        income: Decimal,
        expense: Decimal,
        tax_amount: Decimal
    ) -> str:
        """
        Генерация декларации УСН

        Returns:
            Путь к сохраненному файлу
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Декларация УСН"

        # Заголовок
        ws['A1'] = "НАЛОГОВАЯ ДЕКЛАРАЦИЯ"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = "по налогу, уплачиваемому в связи с применением УСН"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A3'] = f"за {year} год"

        row = 5
        ws[f'A{row}'] = f"Налогоплательщик: {settings.COMPANY_NAME}"
        row += 1
        ws[f'A{row}'] = f"ИНН: {settings.COMPANY_INN}"
        row += 1
        ws[f'A{row}'] = "Объект налогообложения: Доходы, уменьшенные на величину расходов"
        row += 2

        # Раздел 2.2: Расчет налога
        ws.cell(row=row, column=1, value="Раздел 2.2. Расчет налога")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        row += 2

        ws.cell(row=row, column=1, value="Показатель")
        ws.cell(row=row, column=2, value="Сумма (руб.)")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).font = Font(bold=True)
        row += 1

        # Доходы
        ws.cell(row=row, column=1, value="Доходы (210)")
        ws.cell(row=row, column=2, value=float(income))
        row += 1

        # Расходы
        ws.cell(row=row, column=1, value="Расходы (220)")
        ws.cell(row=row, column=2, value=float(expense))
        row += 1

        # Налоговая база
        tax_base = income - expense
        ws.cell(row=row, column=1, value="Налоговая база (240-250)")
        ws.cell(row=row, column=2, value=float(tax_base))
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        # Ставка
        ws.cell(row=row, column=1, value="Ставка налога (%)")
        ws.cell(row=row, column=2, value=15.0)
        row += 1

        # Сумма налога
        ws.cell(row=row, column=1, value="Сумма налога (270)")
        ws.cell(row=row, column=2, value=float(tax_amount))
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        ws.cell(row=row, column=2).font = Font(bold=True, size=11)
        row += 2

        # Минимальный налог
        min_tax = (income * Decimal('0.01')).quantize(Decimal('0.01'))
        ws.cell(row=row, column=1, value="Минимальный налог 1% (280)")
        ws.cell(row=row, column=2, value=float(min_tax))
        row += 2

        # К уплате
        final_tax = max(tax_amount, min_tax)
        ws.cell(row=row, column=1, value="ИТОГО К УПЛАТЕ:")
        ws.cell(row=row, column=2, value=float(final_tax))
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=2).font = Font(bold=True, size=12)

        # Сохранение
        filename = f"USN_Declaration_{year}_{date.today().isoformat()}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)

        logger.info(f"USN Declaration generated: {filepath}")
        return str(filepath)
